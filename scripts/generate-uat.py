"""
Generates Tuscany Storage UAT (User Acceptance Testing) workbook.
Industry-standard format: cover, instructions, environment, test cases,
defect log, summary, sign-off.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import date

OUT = r"D:\Users\silvi\Documents\Hive\Clients\Tuscany\tuscany-storage\docs\Tuscany_Storage_UAT.xlsx"

# ----- Styles -----
NAVY = "1F3A5F"
LIGHT_BLUE = "DCE6F1"
GREY = "D9D9D9"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFEB9C"

thin = Side(border_style="thin", color="808080")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(name="Calibri", size=18, bold=True, color=WHITE)
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=12, bold=True, color=NAVY)
body_font = Font(name="Calibri", size=10)
body_bold = Font(name="Calibri", size=10, bold=True)
small = Font(name="Calibri", size=9, italic=True, color="595959")

navy_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=NAVY)
section_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
alt_fill = PatternFill("solid", fgColor=LIGHT_GREY)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers, fill=header_fill, font=header_font, height=30):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = center
        c.border = border
    ws.row_dimensions[row].height = height


def apply_body_style(ws, start_row, end_row, n_cols, alt=True):
    for r in range(start_row, end_row + 1):
        fill = alt_fill if (alt and (r - start_row) % 2 == 1) else None
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = left_wrap
            cell.font = body_font
            cell.border = border
            if fill:
                cell.fill = fill


wb = Workbook()

# ============================================================
# 1. COVER PAGE
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False

ws.merge_cells("B2:G4")
c = ws["B2"]
c.value = "TUSCANY VILLAGE SELF STORAGE"
c.font = Font(name="Calibri", size=22, bold=True, color=NAVY)
c.alignment = center

ws.merge_cells("B5:G6")
c = ws["B5"]
c.value = "User Acceptance Testing (UAT) Plan & Test Cases"
c.font = Font(name="Calibri", size=14, color="404040")
c.alignment = center

ws.merge_cells("B8:G8")
c = ws["B8"]
c.value = "Project Information"
c.font = title_font
c.fill = navy_fill
c.alignment = center
ws.row_dimensions[8].height = 28

meta = [
    ("Project Name", "Tuscany Village Self Storage — Management Platform"),
    ("Document Type", "User Acceptance Testing (UAT) Plan"),
    ("Document Version", "1.0"),
    ("Release / Build", "Production candidate"),
    ("Prepared By", "Hive Media Studio"),
    ("Prepared For", "David (Tuscany Village Self Storage)"),
    ("Date Issued", date.today().isoformat()),
    ("Production URL", "https://tuscany-storage.onrender.com"),
    ("Repository", "github.com/Shiru04/tuscany-storage"),
    ("Test Environment", "Staging — production-equivalent data"),
    ("Browser Support", "Chrome (latest), Safari (latest), Edge (latest), Firefox (latest)"),
    ("Mobile Support", "iOS Safari, Android Chrome (responsive)"),
]
for i, (k, v) in enumerate(meta):
    r = 10 + i
    ws.cell(row=r, column=2, value=k).font = body_bold
    ws.cell(row=r, column=2).fill = section_fill
    ws.cell(row=r, column=2).border = border
    ws.cell(row=r, column=2).alignment = left_wrap
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=3, value=v)
    cell.font = body_font
    cell.alignment = left_wrap
    cell.border = border
    ws.row_dimensions[r].height = 20

# Approval section
ar = 10 + len(meta) + 2
ws.merge_cells(start_row=ar, start_column=2, end_row=ar, end_column=7)
c = ws.cell(row=ar, column=2, value="Final Acceptance Sign-Off")
c.font = title_font
c.fill = navy_fill
c.alignment = center
ws.row_dimensions[ar].height = 28

signoff_headers = ["Role", "Name", "Signature", "Date"]
header_row(ws, ar + 1, [""] + signoff_headers, height=24)
ws.cell(row=ar + 1, column=1).fill = navy_fill
ws.cell(row=ar + 1, column=1).border = border
roles = ["Client / Product Owner", "Project Manager", "QA Lead", "Technical Lead"]
for i, role in enumerate(roles):
    r = ar + 2 + i
    ws.cell(row=r, column=2, value=role).font = body_bold
    ws.cell(row=r, column=2).fill = section_fill
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).alignment = left_wrap
    ws.row_dimensions[r].height = 40

set_col_widths(ws, [3, 26, 22, 22, 22, 22, 22])

# ============================================================
# 2. INSTRUCTIONS
# ============================================================
ws = wb.create_sheet("Instructions")
ws.sheet_view.showGridLines = False
ws.merge_cells("B2:F2")
c = ws["B2"]
c.value = "How to Use This Document"
c.font = title_font
c.fill = navy_fill
c.alignment = center
ws.row_dimensions[2].height = 28

intro = (
    "This workbook contains the User Acceptance Test (UAT) cases for Tuscany Village Self Storage. "
    "Its purpose is to verify, prior to production release, that every user-facing feature behaves as expected "
    "from the perspective of the end user (Admin, Tenant, Public visitor). Each test case below describes a scenario, "
    "the exact steps to reproduce, and the expected outcome. The tester executes each case and records the result."
)
ws.merge_cells("B4:F7")
c = ws["B4"]
c.value = intro
c.font = body_font
c.alignment = left_wrap

# Workflow
ws.merge_cells("B9:F9")
c = ws["B9"]
c.value = "Testing Workflow"
c.font = section_font
c.fill = section_fill
c.alignment = left_wrap

workflow = [
    "1. Open the 'Test Environment' tab and confirm test credentials and URLs are accessible.",
    "2. Open the 'Test Cases' tab. Filter by module if you want to focus on one area.",
    "3. For each test case: read Preconditions, execute the Steps, compare against Expected Result.",
    "4. Record what actually happened in 'Actual Result' and select a Status from the dropdown.",
    "5. If Status = Fail or Blocked: open the 'Defect Log' tab and create an entry. Reference its Defect ID back in the test case row.",
    "6. Fill in Tester Name and Test Date for every case you execute.",
    "7. The 'Summary' tab automatically counts results — no manual update needed.",
    "8. When all Critical and High cases pass, parties sign the 'Cover' tab to release the build to production.",
]
for i, line in enumerate(workflow):
    r = 10 + i
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2, value=line).font = body_font
    ws.cell(row=r, column=2).alignment = left_wrap

# Status definitions
sr = 10 + len(workflow) + 2
ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=6)
c = ws.cell(row=sr, column=2, value="Status Definitions")
c.font = section_font
c.fill = section_fill

statuses = [
    ("Not Run", "Test has not been executed yet (default state)."),
    ("Pass", "Actual result matches expected result. No defect."),
    ("Fail", "Actual result does not match expected result. A defect must be raised."),
    ("Blocked", "Test cannot be executed because of an environmental issue or dependency on a failed earlier test."),
    ("N/A", "Test is not applicable to this release (feature disabled, deferred, etc.). Add reason in Comments."),
]
sr2 = sr + 1
ws.cell(row=sr2, column=2, value="Status").font = body_bold
ws.cell(row=sr2, column=2).fill = section_fill
ws.merge_cells(start_row=sr2, start_column=3, end_row=sr2, end_column=6)
ws.cell(row=sr2, column=3, value="Meaning").font = body_bold
ws.cell(row=sr2, column=3).fill = section_fill
for i, (s, d) in enumerate(statuses):
    r = sr2 + 1 + i
    ws.cell(row=r, column=2, value=s).font = body_bold
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=d).alignment = left_wrap
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = border

# Severity definitions
sevr = sr2 + len(statuses) + 3
ws.merge_cells(start_row=sevr, start_column=2, end_row=sevr, end_column=6)
c = ws.cell(row=sevr, column=2, value="Defect Severity Definitions")
c.font = section_font
c.fill = section_fill

severities = [
    ("Critical", "Blocks core business operation. No workaround. Examples: cannot log in, payments fail across the board, gate access broken for all tenants."),
    ("High", "Major function broken or significantly degraded. Workaround exists but is painful. Must be fixed before release."),
    ("Medium", "Feature partially works or has noticeable defect that does not block primary flow. Should be fixed soon."),
    ("Low", "Cosmetic issue, minor inconvenience, or edge case. Can be deferred."),
]
sevr2 = sevr + 1
ws.cell(row=sevr2, column=2, value="Severity").font = body_bold
ws.cell(row=sevr2, column=2).fill = section_fill
ws.merge_cells(start_row=sevr2, start_column=3, end_row=sevr2, end_column=6)
ws.cell(row=sevr2, column=3, value="Meaning").font = body_bold
ws.cell(row=sevr2, column=3).fill = section_fill
for i, (s, d) in enumerate(severities):
    r = sevr2 + 1 + i
    ws.cell(row=r, column=2, value=s).font = body_bold
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=d).alignment = left_wrap
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = border

# Priority definitions
pr = sevr2 + len(severities) + 3
ws.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=6)
c = ws.cell(row=pr, column=2, value="Test Case Priority")
c.font = section_font
c.fill = section_fill

priorities = [
    ("Critical", "Must pass for release. Covers core business functions (auth, payment, gate access, move-in)."),
    ("High", "Must pass for release. Covers major features users rely on daily."),
    ("Medium", "Should pass. Covers secondary flows. Failures discussed before release."),
    ("Low", "Nice to have. Edge cases, cosmetics, rare flows."),
]
pr2 = pr + 1
ws.cell(row=pr2, column=2, value="Priority").font = body_bold
ws.cell(row=pr2, column=2).fill = section_fill
ws.merge_cells(start_row=pr2, start_column=3, end_row=pr2, end_column=6)
ws.cell(row=pr2, column=3, value="Meaning").font = body_bold
ws.cell(row=pr2, column=3).fill = section_fill
for i, (s, d) in enumerate(priorities):
    r = pr2 + 1 + i
    ws.cell(row=r, column=2, value=s).font = body_bold
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=d).alignment = left_wrap
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = border

# Exit criteria
er = pr2 + len(priorities) + 3
ws.merge_cells(start_row=er, start_column=2, end_row=er, end_column=6)
c = ws.cell(row=er, column=2, value="Release / Sign-Off Criteria")
c.font = section_font
c.fill = section_fill
exit_criteria = [
    "100% of Critical-priority test cases have Status = Pass.",
    "≥ 95% of High-priority test cases have Status = Pass.",
    "No open defects with Severity = Critical or High.",
    "All Critical / High defects raised during testing are either Fixed-Verified or Accepted-with-Workaround.",
    "All sign-off rows on the Cover sheet are completed.",
]
for i, line in enumerate(exit_criteria):
    r = er + 1 + i
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2, value="• " + line).alignment = left_wrap

set_col_widths(ws, [2, 18, 25, 25, 25, 30])

# ============================================================
# 3. TEST ENVIRONMENT
# ============================================================
ws = wb.create_sheet("Test Environment")
ws.sheet_view.showGridLines = False
ws.merge_cells("B2:F2")
c = ws["B2"]
c.value = "Test Environment & Credentials"
c.font = title_font
c.fill = navy_fill
c.alignment = center
ws.row_dimensions[2].height = 28

ws.merge_cells("B4:F4")
ws["B4"].value = "URLs"
ws["B4"].font = section_font
ws["B4"].fill = section_fill

urls = [
    ("Production", "https://tuscany-storage.onrender.com"),
    ("Admin Login", "https://tuscany-storage.onrender.com/login"),
    ("Tenant Portal", "https://tuscany-storage.onrender.com/portal"),
    ("Public Site", "https://tuscany-storage.onrender.com"),
    ("Reserve Unit (Move-In Funnel)", "https://tuscany-storage.onrender.com/reserve"),
]
for i, (k, v) in enumerate(urls):
    r = 5 + i
    ws.cell(row=r, column=2, value=k).font = body_bold
    ws.cell(row=r, column=2).fill = alt_fill
    ws.cell(row=r, column=2).border = border
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(row=r, column=3, value=v).font = body_font
    ws.cell(row=r, column=3).border = border

# Credentials
cr = 5 + len(urls) + 2
ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=6)
c = ws.cell(row=cr, column=2, value="Test Credentials")
c.font = section_font
c.fill = section_fill

ws.cell(row=cr + 1, column=2, value="Role").font = body_bold
ws.cell(row=cr + 1, column=3, value="Email / Username").font = body_bold
ws.cell(row=cr + 1, column=4, value="Password").font = body_bold
ws.merge_cells(start_row=cr + 1, start_column=5, end_row=cr + 1, end_column=6)
ws.cell(row=cr + 1, column=5, value="Notes").font = body_bold
for col in range(2, 7):
    ws.cell(row=cr + 1, column=col).fill = section_fill
    ws.cell(row=cr + 1, column=col).border = border

creds = [
    ("Admin", "admin@tuscanystorage.com", "(provided separately)", "Full admin access"),
    ("Tenant 1", "john@example.com", "tenant123", "Active tenant with unit + autopay on"),
    ("Tenant 2", "jane@example.com", "tenant123", "Active tenant with balance owed"),
    ("Tenant 3", "bob@example.com", "tenant123", "Delinquent tenant (for lockout tests)"),
]
for i, row in enumerate(creds):
    r = cr + 2 + i
    for col, val in enumerate(row[:3], 2):
        ws.cell(row=r, column=col, value=val).font = body_font
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).alignment = left_wrap
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5, value=row[3]).font = body_font
    ws.cell(row=r, column=5).border = border
    ws.cell(row=r, column=5).alignment = left_wrap

# Stripe test cards
sc = cr + 2 + len(creds) + 2
ws.merge_cells(start_row=sc, start_column=2, end_row=sc, end_column=6)
c = ws.cell(row=sc, column=2, value="Stripe Test Cards (TEST mode only)")
c.font = section_font
c.fill = section_fill

ws.cell(row=sc + 1, column=2, value="Card Number").font = body_bold
ws.cell(row=sc + 1, column=3, value="Expected Behaviour").font = body_bold
ws.merge_cells(start_row=sc + 1, start_column=4, end_row=sc + 1, end_column=6)
ws.cell(row=sc + 1, column=4, value="Notes (use any future expiry, any CVC, any ZIP)").font = body_bold
for col in range(2, 7):
    ws.cell(row=sc + 1, column=col).fill = section_fill
    ws.cell(row=sc + 1, column=col).border = border

cards = [
    ("4242 4242 4242 4242", "Successful charge", "Default success card."),
    ("4000 0025 0000 3155", "Requires 3D Secure", "Tests Strong Customer Authentication challenge."),
    ("4000 0000 0000 9995", "Decline — insufficient funds", "Used for decline-handling tests."),
    ("4000 0000 0000 0002", "Decline — generic", "Generic decline."),
    ("4000 0000 0000 0341", "Attach succeeds, future charge fails", "Use for autopay failure tests."),
]
for i, row in enumerate(cards):
    r = sc + 2 + i
    ws.cell(row=r, column=2, value=row[0]).font = body_font
    ws.cell(row=r, column=2).border = border
    ws.cell(row=r, column=3, value=row[1]).font = body_font
    ws.cell(row=r, column=3).border = border
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(row=r, column=4, value=row[2]).font = body_font
    ws.cell(row=r, column=4).border = border
    ws.cell(row=r, column=4).alignment = left_wrap

# Prerequisites
pre = sc + 2 + len(cards) + 2
ws.merge_cells(start_row=pre, start_column=2, end_row=pre, end_column=6)
c = ws.cell(row=pre, column=2, value="Tester Prerequisites")
c.font = section_font
c.fill = section_fill
prereqs = [
    "Working email inbox for receiving test confirmations and receipts (e.g., +uat alias on a real address).",
    "Mobile phone capable of receiving SMS (for text-to-open and SMS notification tests).",
    "Modern browser (Chrome / Edge / Safari latest).",
    "Access to the staging environment URL above.",
    "Stripe test cards listed in this sheet — do NOT use real credit cards.",
    "(For PDK / gate tests) coordination with the on-site PDK device, or use of the sandbox simulator.",
]
for i, p in enumerate(prereqs):
    r = pre + 1 + i
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2, value="• " + p).font = body_font
    ws.cell(row=r, column=2).alignment = left_wrap

set_col_widths(ws, [2, 28, 32, 18, 22, 22])

# ============================================================
# 4. TEST CASES
# ============================================================
ws = wb.create_sheet("Test Cases")
ws.freeze_panes = "A2"

tc_headers = [
    "Test ID", "Module", "Sub-Module", "Priority", "Test Title",
    "Preconditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Status", "Severity", "Tester", "Test Date",
    "Comments", "Defect ID",
]
header_row(ws, 1, tc_headers, height=36)

# ---- TEST CASE DEFINITIONS ----
# Each row: (module, sub_module, priority, title, precondition, steps, test_data, expected)
TC = []

# ===== AUTHENTICATION =====
TC += [
    ("Authentication", "Login", "Critical",
     "Admin can log in with valid credentials",
     "Admin account exists.",
     "1. Navigate to /login.\n2. Enter admin email and password.\n3. Click Sign In.",
     "admin@tuscanystorage.com / valid password",
     "User is redirected to /admin dashboard. Sidebar nav is visible. Session is established."),
    ("Authentication", "Login", "Critical",
     "Tenant can log in and lands on portal",
     "Tenant account exists with assigned unit.",
     "1. Navigate to /login.\n2. Enter tenant email + password.\n3. Click Sign In.",
     "john@example.com / tenant123",
     "User is redirected to /portal. Dashboard shows tenant's name, balance, and rentals."),
    ("Authentication", "Login", "High",
     "Login fails with wrong password",
     "Tenant account exists.",
     "1. Navigate to /login.\n2. Enter valid email + wrong password.\n3. Click Sign In.",
     "john@example.com / WrongPass",
     "Error message displayed (generic 'invalid credentials'). User stays on login page. No session created."),
    ("Authentication", "Login", "Medium",
     "Login fails with non-existent email",
     "—",
     "1. Navigate to /login.\n2. Enter random email + any password.\n3. Click Sign In.",
     "noone@example.com / anything",
     "Same generic error as wrong password (no enumeration of valid emails)."),
    ("Authentication", "Password reset", "High",
     "Tenant can request password reset",
     "Tenant account exists.",
     "1. Click 'Forgot Password' on login screen.\n2. Enter tenant email.\n3. Click Send.",
     "john@example.com",
     "Confirmation screen shown. Reset email arrives at inbox within 2 minutes with reset link."),
    ("Authentication", "Password reset", "High",
     "Reset link sets a new password",
     "Reset email received within last 24h.",
     "1. Open the reset link.\n2. Enter new password (meets strength rules) twice.\n3. Submit.\n4. Log in with the new password.",
     "Strong new password (e.g., Tuscany!2026)",
     "Password is updated. Login with new password succeeds. Login with old password fails."),
    ("Authentication", "Password reset", "Medium",
     "Reset link expires after use",
     "Password just reset using a link.",
     "1. Open the same reset link a second time.",
     "—",
     "Link is rejected. Page shows 'link expired/invalid' message."),
    ("Authentication", "Logout", "High",
     "User can log out and session ends",
     "Logged in as any role.",
     "1. Click user menu in top-right.\n2. Click Sign Out.\n3. Try to navigate to /admin or /portal.",
     "—",
     "Redirected to /login. Protected pages redirect to login when accessed."),
    ("Authentication", "Session", "Medium",
     "Admin route blocks tenant role",
     "Logged in as tenant.",
     "1. While logged in as tenant, navigate directly to /admin in URL bar.",
     "—",
     "Tenant is redirected away (to /portal or /login). No admin content rendered."),
]

# ===== PUBLIC SITE =====
TC += [
    ("Public Site", "Home", "Medium",
     "Public homepage loads and displays facility info",
     "—",
     "1. Open the production URL in a fresh browser.",
     "—",
     "Homepage loads. Facility name, hero copy, and CTA buttons are visible. No JS errors in console."),
    ("Public Site", "Units listing", "High",
     "Available units are displayed publicly",
     "At least one unit has status = available.",
     "1. From homepage, click 'Units' or 'Reserve'.\n2. Browse the unit grid.",
     "—",
     "Units shown with size, monthly price, features. Filters by size work. Sold-out / occupied units do not appear as available."),
    ("Public Site", "Size guide", "Low",
     "Size guide page renders",
     "—",
     "1. Navigate to /size-guide.",
     "—",
     "Visual comparisons for 5x5, 5x10, 10x10, 10x15, 10x20 displayed. Page is responsive on mobile."),
    ("Public Site", "Contact form", "High",
     "Visitor can submit contact form",
     "—",
     "1. Navigate to /contact.\n2. Fill name, email, subject, message.\n3. Submit.",
     "Use a real reachable email",
     "Success confirmation displayed. Admin receives notification email. Submission appears in /admin/site-forms."),
    ("Public Site", "Waiting list", "High",
     "Visitor can join waiting list for a size",
     "—",
     "1. Navigate to /waiting-list.\n2. Fill form (name, email, phone, requested size, notes).\n3. Complete captcha.\n4. Submit.",
     "Valid contact info",
     "Confirmation shown. Entry appears in /admin/waiting-list. Confirmation email received."),
    ("Public Site", "Blog", "Low",
     "Blog posts list and detail render",
     "At least one post is published.",
     "1. Navigate to /blog.\n2. Click a post.",
     "—",
     "Posts listed with title and excerpt. Detail page shows full post content."),
]

# ===== MOVE-IN =====
TC += [
    ("Move-In", "Funnel", "Critical",
     "New tenant completes full move-in (happy path)",
     "Unit available. Tester has a working email + Stripe test card.",
     "1. From /units pick a unit, click Reserve.\n2. Step 1 contact info: fill name/email/phone/address + SMS consent.\n3. Step 2 personal info: DL, employer, emergency contact, etc.\n4. Step 3 alternate contact (optional).\n5. Step 4 billing info: Stripe card.\n6. Step 5 review agreement, e-sign, accept terms.\n7. Submit.",
     "Test card 4242 4242 4242 4242 / 12/30 / 123 / 12345",
     "Reservation succeeds. Tenant + Lease are created. Gate code is auto-assigned. Welcome email with login credentials received. Tenant is redirected to /portal with welcome modal."),
    ("Move-In", "Validation", "High",
     "Move-in form blocks missing required fields",
     "—",
     "1. Start move-in flow.\n2. On Step 1, click Next without filling any required field.",
     "—",
     "Form does not advance. Each required field shows an inline validation message."),
    ("Move-In", "Validation", "High",
     "Move-in form rejects invalid email format",
     "—",
     "1. On Step 1, fill name + phone, enter 'not-an-email' in email field.\n2. Try to advance.",
     "—",
     "Email field shows format-error. Cannot advance."),
    ("Move-In", "Payment", "Critical",
     "Move-in is rejected when card is declined",
     "Unit available.",
     "1. Complete steps 1–4 of move-in.\n2. Use decline test card on billing step.\n3. Submit.",
     "Test card 4000 0000 0000 0002",
     "Payment fails with clear error. Tenant is NOT created. Lease is NOT created. No gate code issued. User can retry with different card."),
    ("Move-In", "Agreement", "High",
     "Tenant can e-sign agreement during move-in",
     "On the e-sign step of move-in flow.",
     "1. Draw or type signature.\n2. Tick the 'I agree' checkbox.\n3. Submit.",
     "—",
     "Signature is captured. Agreement is marked signed. PDF of signed agreement available in tenant portal."),
    ("Move-In", "Reservation fee", "Medium",
     "Reservation fee is charged when configured",
     "Reservation fee for that unit type is configured > $0 in Settings.",
     "1. Complete a move-in for a unit type that has a reservation fee.",
     "Reservation fee = $25 (example)",
     "Reservation fee appears in pricing summary on the final step. Tenant is charged at submission (or added to first invoice, per facility setting)."),
    ("Move-In", "Welcome", "Medium",
     "Welcome modal appears on tenant's first portal visit",
     "Tenant just completed move-in.",
     "1. After move-in submission, observe portal first-load.",
     "—",
     "Modal appears with welcome message and instructions. Modal can be dismissed and does not re-appear on next login."),
]

# ===== TENANT PORTAL =====
TC += [
    ("Tenant Portal", "Dashboard", "Critical",
     "Tenant dashboard shows correct balance and rentals",
     "Tenant is logged in with active lease and known balance.",
     "1. Navigate to /portal.",
     "—",
     "Dashboard displays tenant name, unit(s) with size/rate/next-bill-date, current balance, next-bill date, deposit, agreement-signed status."),
    ("Tenant Portal", "Profile", "High",
     "Tenant can update profile when enabled",
     "Setting customersCanEditProfile = true. Logged in as tenant.",
     "1. Navigate to /portal/profile.\n2. Change phone number.\n3. Save.",
     "Any new valid phone number",
     "Update succeeds. Reload shows new value. Admin tenant detail page reflects the new phone."),
    ("Tenant Portal", "Profile", "Medium",
     "Profile edit hidden when disabled",
     "Setting customersCanEditProfile = false.",
     "1. Navigate to /portal/profile as tenant.",
     "—",
     "Profile fields are read-only. No Save button visible."),
    ("Tenant Portal", "Billing", "High",
     "Tenant can add a payment method",
     "Tenant has no card on file.",
     "1. Navigate to /portal/billing.\n2. Click Add Card.\n3. Enter Stripe test card.\n4. Save.",
     "Test card 4242 4242 4242 4242",
     "Card is saved. Last 4 digits and expiry are displayed on /portal/billing. Stripe Dashboard reflects new payment method on this customer."),
    ("Tenant Portal", "Billing", "High",
     "Tenant can replace existing card",
     "Tenant already has a card on file.",
     "1. Navigate to /portal/billing.\n2. Click Update Card.\n3. Enter new test card.\n4. Save.",
     "Test card 4000 0025 0000 3155",
     "Old card is replaced. New card last 4 / expiry are displayed. Stripe customer's default payment method is updated."),
    ("Tenant Portal", "Billing", "High",
     "Tenant can toggle autopay on/off",
     "Logged in as tenant with card on file.",
     "1. Navigate to /portal/billing.\n2. Toggle autopay on (then off).\n3. Reload.",
     "—",
     "Toggle persists. Admin tenant page shows correct autopay status."),
    ("Tenant Portal", "Payments", "Critical",
     "Tenant can make a one-time payment",
     "Tenant has an outstanding balance and a saved card.",
     "1. Navigate to /portal/payments/new.\n2. Choose 'Pay full balance' (or custom amount).\n3. Select saved card.\n4. Submit.",
     "—",
     "Payment succeeds. Balance updates to $0 (or reduced). Receipt email is received. Payment appears in history."),
    ("Tenant Portal", "Payments", "High",
     "Payment with 3D Secure card triggers SCA challenge",
     "Tenant has outstanding balance.",
     "1. Make a payment using the 3DS-required test card.",
     "Card 4000 0025 0000 3155",
     "3D Secure popup/redirect appears. After completing challenge, payment succeeds."),
    ("Tenant Portal", "Payments", "High",
     "Failed payment is recorded and surfaces to tenant",
     "Tenant has outstanding balance.",
     "1. Make a payment using the decline card.",
     "Card 4000 0000 0000 0002",
     "Clear error message shown. Balance does not change. Failed payment recorded in admin payments report."),
    ("Tenant Portal", "Payments", "Medium",
     "Payment history shows all completed payments",
     "Tenant has at least 2 historical payments.",
     "1. Navigate to /portal/payments.",
     "—",
     "Payment history table lists date, amount, status, and link to receipt for each payment."),
    ("Tenant Portal", "Gate code", "Critical",
     "Tenant can view their gate code",
     "Logged in as tenant with gate code assigned.",
     "1. Navigate to /portal/gate-code.\n2. Click Reveal.",
     "—",
     "Code is blurred by default. Reveal toggle un-blurs the code. Same code shown as in admin tenant detail."),
    ("Tenant Portal", "Gate code", "High",
     "Tenant can request a new gate code",
     "Tenant logged in.",
     "1. Navigate to /portal/gate-code.\n2. Click 'Request New Code'.\n3. Confirm.",
     "—",
     "New code generated and displayed. SMS with new code received within 2 minutes. Old code no longer works at gate (verify if PDK is live)."),
    ("Tenant Portal", "Gate access history", "Medium",
     "Tenant sees their own access events",
     "Tenant has at least one historical gate entry.",
     "1. Navigate to /portal/gate-code.\n2. Scroll to access history.",
     "—",
     "List shows tenant's entry/exit events with timestamp, gate name, source. Filterable by event type and date range."),
    ("Tenant Portal", "Move-out request", "High",
     "Tenant can request a move-out date",
     "Logged in as tenant with active lease, setting customersCanScheduleMoveOuts = true.",
     "1. Navigate to /portal/move-out.\n2. Pick lease, date, optional notes.\n3. Submit.",
     "Move-out date 7 days from today",
     "Confirmation displayed. Admin sees the request in /admin/move-out as 'pending_moveout'."),
    ("Tenant Portal", "Move-out request", "Medium",
     "Move-out nav item hidden when disabled",
     "Setting customersCanScheduleMoveOuts = false.",
     "1. Log in as tenant.\n2. Inspect portal sidebar.",
     "—",
     "'Move Out' item does not appear in the sidebar. Direct URL /portal/move-out returns a friendly disabled message or 404."),
    ("Tenant Portal", "Lease", "High",
     "Tenant can sign an unsigned lease from portal",
     "Tenant has lease with status active but unsigned.",
     "1. Navigate to /portal/lease/sign?leaseId=...\n2. Draw / type signature.\n3. Submit.",
     "—",
     "Signature captured. Lease marked signed. Signed PDF available in /portal/lease and emailed to tenant."),
    ("Tenant Portal", "Instructions", "Low",
     "Instructions page renders configured content",
     "Settings.newRenterInstructions has content.",
     "1. Navigate to /portal/instructions.",
     "—",
     "Page shows the formatted instructions from settings."),
]

# ===== ADMIN: TENANTS =====
TC += [
    ("Admin / Tenants", "List", "High",
     "Admin sees full tenant list with key columns",
     "Database has tenants.",
     "1. Log in as admin.\n2. Navigate to /admin/tenants.",
     "—",
     "Table loads with name, unit, status, balance, autopay indicator. Pagination works. Search by name returns matches."),
    ("Admin / Tenants", "Search & filter", "Medium",
     "Filter by status returns correct subset",
     "Tenants in multiple statuses exist.",
     "1. On /admin/tenants apply status filter = 'Delinquent'.",
     "—",
     "Only delinquent tenants are listed. Counter updates."),
    ("Admin / Tenants", "Detail", "High",
     "Tenant detail shows contact, leases, balance, billing history",
     "Open a known tenant.",
     "1. Click a tenant row.",
     "—",
     "Detail page shows contact info, active leases, balance, recent 10 transactions, autopay status, default card."),
    ("Admin / Tenants", "Create", "High",
     "Admin can create a new tenant manually",
     "—",
     "1. Click 'New Tenant' on /admin/tenants.\n2. Fill all required fields.\n3. Save.",
     "Unique email, name, phone",
     "Tenant is created. Confirmation shown. New tenant appears in list. Optional welcome email is sent if enabled."),
    ("Admin / Tenants", "Edit", "Medium",
     "Admin can edit tenant contact info",
     "Open a tenant detail.",
     "1. Click Edit.\n2. Change address.\n3. Save.",
     "—",
     "Address updated. Reload confirms change."),
    ("Admin / Tenants", "Password reset", "Medium",
     "Admin can send password reset link to tenant",
     "Open a tenant.",
     "1. Click 'Send Password Reset'.",
     "—",
     "Confirmation shown. Tenant inbox receives reset email."),
    ("Admin / Tenants", "Rent unit", "Critical",
     "Admin can rent a unit to an existing tenant",
     "Tenant with no active lease exists. Unit with status=available exists.",
     "1. Open tenant.\n2. Click Rent Unit.\n3. Select unit.\n4. Configure rate, deposit, proration, billing day.\n5. Submit.",
     "Available unit, monthly rate $150, billing day 1",
     "Lease is created. Unit status changes to occupied. First invoice generated. Tenant appears with new unit on portal."),
    ("Admin / Tenants", "Reserve unit", "Medium",
     "Admin can reserve a unit for a tenant",
     "Unit available, reservation fee may be configured.",
     "1. Open tenant.\n2. Reserve Unit → pick unit → set reserve end date → save.",
     "Reserve until 7 days from today",
     "Unit changes to 'reserved'. Reservation fee (if configured) is added to tenant balance."),
    ("Admin / Tenants", "Gate access", "Critical",
     "Admin can regenerate a tenant gate code",
     "Tenant has unit + existing code.",
     "1. Open tenant Gate Access.\n2. Click Regenerate Code.",
     "—",
     "Code changes. Tenant receives SMS notification with new code (if SMS enabled). Old code stops working at gate."),
    ("Admin / Tenants", "Gate access", "High",
     "Admin can manually lock out a tenant",
     "Tenant has active gate access.",
     "1. Open tenant Gate Access.\n2. Set lockout = today, add reason.\n3. Save.",
     "Reason: 'UAT — manual lockout'",
     "Tenant status updates to 'Locked Out'. Tenant removed from PDK access group (when PDK enabled). Tenant portal shows locked-out banner."),
    ("Admin / Tenants", "Gate access", "High",
     "Admin can unlock a tenant",
     "Tenant is currently locked out.",
     "1. Open tenant Gate Access.\n2. Click Unlock / clear lockout.",
     "—",
     "Status returns to Active. Gate access restored at PDK. Tenant can use code again."),
    ("Admin / Tenants", "Make payment", "Critical",
     "Admin can charge tenant card off-session",
     "Tenant has card on file.",
     "1. Open tenant.\n2. Click Make Payment.\n3. Enter amount + description.\n4. Submit.",
     "Amount $50, type rent",
     "Payment succeeds via Stripe. Receipt issued. Balance adjusts. Charge appears in tenant billing history."),
    ("Admin / Tenants", "Add credit", "Medium",
     "Admin can apply a credit to tenant account",
     "Open tenant.",
     "1. Click Add Credit.\n2. Enter amount + reason.\n3. Save.",
     "Credit $20, reason 'UAT credit'",
     "Negative charge (credit) appears in billing history. Balance reduces by amount."),
    ("Admin / Tenants", "Refund", "High",
     "Admin can refund a card payment",
     "Tenant has a recent successful card payment.",
     "1. Open tenant billing history.\n2. Click a card payment → Refund.\n3. Enter amount + reason.\n4. Submit.",
     "Refund $10 of a $50 payment",
     "Stripe refund initiated (verify in Stripe Dashboard). Refund record created. Tenant balance reflects refund."),
    ("Admin / Tenants", "Recurring fees", "Medium",
     "Admin can add a recurring monthly fee to a tenant",
     "Open tenant Recurring Fees page.",
     "1. New recurring fee → name, amount, due day, auto-charge flag.\n2. Save.",
     "$10 Insurance, day 1, auto-charge on",
     "Fee appears in tenant recurring fees list. On the due day, autopay job picks it up and charges if card available."),
    ("Admin / Tenants", "Letters", "Low",
     "Admin can print a late notice for a tenant",
     "Open tenant.",
     "1. Letters → Late Notice → Print.",
     "—",
     "PDF generated with tenant name, balance, due date, and configured template."),
    ("Admin / Tenants", "Move-out finalize", "Critical",
     "Admin can finalize a move-out and issue receipt",
     "Tenant has pending move-out request.",
     "1. Open /admin/move-out → tenant.\n2. Review final charges.\n3. Click Finalize.\n4. Send receipt via email.",
     "—",
     "Lease ended. Unit returns to available status. Tenant removed from PDK group. Move-out receipt email received with final settlement."),
    ("Admin / Tenants", "Duplicates", "Low",
     "Duplicate detection lists likely duplicates",
     "At least 2 tenant records share surname/address.",
     "1. Navigate to /admin/tenants/duplicates.",
     "—",
     "Likely duplicates listed with confidence score. Merge action consolidates the two records."),
]

# ===== ADMIN: UNITS =====
TC += [
    ("Admin / Units", "List view", "High",
     "Unit list loads with status, type, tenant, rate",
     "Units exist in DB.",
     "1. Navigate to /admin/units/list.",
     "—",
     "Table shows columns and supports sorting by number/rate. Filters by type and status work."),
    ("Admin / Units", "Site map", "Medium",
     "Site map renders units with status colours",
     "Units exist with floor layout.",
     "1. Navigate to /admin/units/site-map.",
     "—",
     "Visual layout displays units coloured by status (available, occupied, reserved, down). Clicking a unit opens its detail."),
    ("Admin / Units", "Create", "High",
     "Admin can create a new unit",
     "—",
     "1. Click New Unit.\n2. Fill number, type, size, floor, price, features.\n3. Save.",
     "Unit 999, 10x10, climate-controlled, $150/mo",
     "Unit is created. Appears in list and on site map."),
    ("Admin / Units", "Edit", "Medium",
     "Admin can edit unit price and features",
     "Open a unit detail.",
     "1. Click Edit.\n2. Change street price.\n3. Save.",
     "New price $175",
     "Update persists. Listing reflects new rate."),
    ("Admin / Units", "Maintenance", "Medium",
     "Admin can mark unit down for maintenance",
     "Open an available unit.",
     "1. Set status = Down.\n2. Save.",
     "—",
     "Unit removed from public availability. Status shown as Down in admin views."),
]

# ===== ADMIN: PAYMENTS =====
TC += [
    ("Admin / Payments", "List", "High",
     "Payments list shows all transactions",
     "Payments exist.",
     "1. Navigate to /admin/payments.",
     "—",
     "Table lists payments with date, tenant, type, status, amount. Filters by date range and status work."),
    ("Admin / Payments", "Export", "Medium",
     "Admin can export payments to CSV",
     "Payments list loaded with at least one row.",
     "1. Click Export CSV.",
     "—",
     "CSV download starts. File contains the filtered rows with all visible columns."),
    ("Admin / Payments", "Transaction detail", "Medium",
     "Admin can view full payment record",
     "At least one successful card payment exists.",
     "1. Open a payment row.",
     "—",
     "Detail shows amount, status, Stripe PaymentIntent / Charge ID, line items, and links to refund and resend receipt."),
    ("Admin / Payments", "Resend receipt", "Low",
     "Admin can resend a receipt to tenant",
     "Open a payment detail.",
     "1. Click Resend Receipt.",
     "—",
     "Confirmation displayed. Tenant inbox receives the receipt email."),
]

# ===== ADMIN: DELINQUENCY =====
TC += [
    ("Admin / Delinquency", "Dashboard", "High",
     "Delinquency dashboard summarises overdue accounts",
     "At least one delinquent tenant exists.",
     "1. Navigate to /admin/delinquency.",
     "—",
     "KPIs show counts by stage (Late, Locked Out, Pre-Lien). Table lists tenants with days overdue, balance, last payment, last contact."),
    ("Admin / Delinquency", "Bulk reminder", "Medium",
     "Admin can send reminder email in bulk",
     "Multiple delinquent tenants.",
     "1. Select 2+ tenants → Send Reminder.",
     "—",
     "Confirmation shown. Email logs recorded. Tenants receive reminder emails."),
    ("Admin / Delinquency", "Lockout", "Critical",
     "Admin can lock out a delinquent tenant",
     "Tenant past due, currently active.",
     "1. Select tenant → Issue Lockout.\n2. Confirm.",
     "—",
     "Tenant status changes to Locked Out. Gate access revoked. Tenant receives lockout notification (per template)."),
    ("Admin / Delinquency", "Release", "High",
     "Admin can release a tenant from lockout after payment",
     "Locked-out tenant whose balance is now $0.",
     "1. Select tenant → Release / Unlock.",
     "—",
     "Status returns to Active. Gate access restored. Tenant can sign in and access portal normally."),
    ("Admin / Delinquency", "Lien notice", "High",
     "Admin can issue lien notice",
     "Tenant locked out for configured days.",
     "1. Select tenant → Issue Lien.",
     "—",
     "Tenant moves to Pre-Lien stage. Lien notice sent per template. Audit entry recorded."),
    ("Admin / Delinquency", "Auction schedule", "Medium",
     "Admin can schedule an auction date",
     "Tenant in Lien stage.",
     "1. Select tenant → Schedule Auction → pick date.",
     "Date 14 days out",
     "Auction date recorded. Tenant receives auction notice. Date visible on tenant detail."),
]

# ===== ADMIN: REPORTS =====
TC += [
    ("Admin / Reports", "Hub", "Medium",
     "Reports hub loads all report categories",
     "—",
     "1. Navigate to /admin/reports.",
     "—",
     "All report cards/links are present (Accounting, Customers, Payments, Facility)."),
    ("Admin / Reports", "Rent roll", "High",
     "Rent roll report lists active leases with rates",
     "Active leases exist.",
     "1. Open Rent Roll report.",
     "—",
     "Report shows unit, tenant, monthly rate, billing day, balance. Total at the bottom matches sum."),
    ("Admin / Reports", "Occupancy", "Medium",
     "Occupancy report shows current occupancy %",
     "Units exist in multiple statuses.",
     "1. Open Occupancy report.",
     "—",
     "Occupancy % shown. Breakdown by unit type listed."),
    ("Admin / Reports", "Gate activity", "Critical",
     "Gate activity report lists access events",
     "Recent gate events exist (real or simulated).",
     "1. Open /admin/reports/gate-activity.\n2. Filter date range.",
     "Last 7 days",
     "Events list shows timestamp, tenant or visitor, device, source (keypad, sms, app), result. Visitor pass events show visitor attribution."),
    ("Admin / Reports", "Move-in / move-out", "Medium",
     "Move-in/out report lists recent events",
     "Recent move-ins and move-outs exist.",
     "1. Open the Move In/Out report.",
     "Last 30 days",
     "Move-in rows show tenant, unit, date, lease start. Move-out rows show tenant, unit, date, final balance."),
    ("Admin / Reports", "Failed payments", "High",
     "Failed payments report lists declines",
     "At least one failed Stripe charge exists.",
     "1. Open Failed/Declined Payments report.",
     "—",
     "Declined rows show tenant, attempt date, amount, decline reason."),
    ("Admin / Reports", "Lockouts", "Medium",
     "Lockouts report lists lockout events with approval state",
     "At least one lockout event exists.",
     "1. Open /admin/reports/lock-out.",
     "—",
     "Events list with status (pending/approved/executed) shown. Approval actions function (if approval workflow enabled)."),
    ("Admin / Reports", "CSV export", "Medium",
     "Reports can be exported to CSV",
     "Any report open with data.",
     "1. Click Export CSV.",
     "—",
     "Download starts. File matches the visible rows."),
]

# ===== ADMIN: SETTINGS =====
TC += [
    ("Admin / Settings", "General", "High",
     "Admin can edit general settings",
     "Logged in as admin.",
     "1. Navigate to /admin/settings/general.\n2. Edit timezone or date format.\n3. Save.",
     "Timezone America/New_York",
     "Change persists. Other pages reflect the new locale settings."),
    ("Admin / Settings", "Facility", "High",
     "Admin can edit facility info and access hours",
     "—",
     "1. /admin/settings/facility.\n2. Change access hours and save.",
     "Hours 06:00–22:00",
     "Change persists. PDK group sync (if enabled) re-pushes the new schedule to entry rule."),
    ("Admin / Settings", "Fees", "High",
     "Admin can edit late, NSF, auction fees",
     "—",
     "1. /admin/settings/fees.\n2. Change late fee amount.\n3. Save.",
     "Late fee $20",
     "Change persists. New late fees added by delinquency job use the new amount."),
    ("Admin / Settings", "Rental", "Medium",
     "Admin can edit billing cycle anchor",
     "—",
     "1. /admin/settings/rental.\n2. Change anchor to 'signup day'.\n3. Save.",
     "—",
     "Change persists. New leases use the new anchor for invoice generation."),
    ("Admin / Settings", "Agreement", "Medium",
     "Admin can edit agreement template",
     "—",
     "1. /admin/settings/agreement.\n2. Make a small text change.\n3. Save.",
     "—",
     "Change persists. New move-ins display the updated agreement on the e-sign step."),
    ("Admin / Settings", "Agreement", "Low",
     "Image upload in agreement editor works (R2)",
     "On agreement editor.",
     "1. Insert image → upload a JPG.",
     "Any small JPG / PNG",
     "Image uploads to Cloudflare R2. URL becomes accessible. Image renders in saved agreement."),
    ("Admin / Settings", "Form fields", "Low",
     "Admin can toggle field visibility on signup form",
     "—",
     "1. /admin/settings/form-fields.\n2. Hide 'Driver's License' from signup.\n3. Save.\n4. Open public move-in flow.",
     "—",
     "DL field is no longer required on signup."),
    ("Admin / Settings", "Late/Lien", "High",
     "Admin can configure escalation days",
     "—",
     "1. /admin/settings/late-lien.\n2. Change 'Days past due → Locked out' to 15.\n3. Save.",
     "—",
     "Change persists. Delinquency job uses the new value."),
    ("Admin / Settings", "Communications", "Medium",
     "Admin can change from-email and reply-to",
     "—",
     "1. /admin/communications/settings.\n2. Change from-name.\n3. Save.\n4. Trigger a test email.",
     "—",
     "Updated from-name appears in subsequent emails."),
]

# ===== ADMIN: COMMUNICATIONS =====
TC += [
    ("Admin / Communications", "Templates list", "Medium",
     "Templates page lists system and custom templates",
     "—",
     "1. /admin/communications/templates.",
     "—",
     "All default templates (receipt, late, lockout, lien, etc.) listed with channel toggles."),
    ("Admin / Communications", "Template edit", "High",
     "Admin can edit a template body and subject",
     "Open the Late Notice template.",
     "1. Change subject and body.\n2. Save.\n3. Trigger a late notice.",
     "Use placeholder {{tenantName}}",
     "Change persists. Sent late notice uses the new subject and body, with placeholders expanded."),
    ("Admin / Communications", "Send custom email", "Medium",
     "Admin can send a one-off email to selected tenants",
     "—",
     "1. /admin/communications/send-email.\n2. Select 1-2 tenants.\n3. Compose subject + body.\n4. Send.",
     "Use tester tenants",
     "Confirmation shown. Tenants receive the email."),
    ("Admin / Communications", "Print batches", "Low",
     "Print batches list shows recent jobs",
     "At least one print batch was generated.",
     "1. /admin/communications/print-batches.",
     "—",
     "Batches listed with status, date, recipient count."),
    ("Admin / Communications", "Mailing labels", "Low",
     "Admin can generate mailing labels",
     "—",
     "1. /admin/communications/mailing-labels.\n2. Select filter (e.g., all active).\n3. Generate PDF.",
     "Avery 5160",
     "PDF generated with correctly formatted addresses."),
]

# ===== ADMIN: RATE MANAGEMENT =====
TC += [
    ("Admin / Rate Mgmt", "Suggestions", "Medium",
     "Rate management surfaces unit-type and rental suggestions",
     "Rate management enabled with rules. Units and leases qualifying for change exist.",
     "1. /admin/rate-management.",
     "—",
     "Unit-type suggestions and rental suggestions tables populated based on rules."),
    ("Admin / Rate Mgmt", "Batch submit", "High",
     "Admin can submit a batch of rate changes",
     "Suggestions present.",
     "1. Select 1-2 rentals.\n2. Submit.",
     "—",
     "Batch is created. Tenants receive notifications per configured channels. Batch appears in /admin/rate-management/batches."),
    ("Admin / Rate Mgmt", "Batch execution", "High",
     "Rate execution cron applies changes on the scheduled date",
     "Batch item exists with changeDate = today.",
     "1. Run cron 'rate-execution' (or wait until natural run).",
     "—",
     "Lease monthlyRate updated. Item status = executed. Confirmation email sent."),
    ("Admin / Rate Mgmt", "Cancel item", "Medium",
     "Admin can cancel a pending rate change",
     "Pending rate batch item exists.",
     "1. Open the batch item → Cancel.",
     "—",
     "Item status = cancelled. Lease rate not changed on the previously scheduled date."),
]

# ===== VISITOR ACCESS =====
TC += [
    ("Visitor Access", "Issue", "Critical",
     "Admin can issue a temporary visitor pass",
     "PDK group bootstrap done (or sandbox enabled).",
     "1. /admin/visitor-access.\n2. New pass → name, purpose, duration 1h.\n3. Submit.",
     "Name 'Electrician', purpose 'Panel check', duration 1h",
     "Pass created. PIN displayed exactly once in reveal dialog. Pass listed under active. PDK holder provisioned (when sync enabled)."),
    ("Visitor Access", "PIN visibility", "High",
     "PIN is shown only once and not in subsequent lists",
     "Pass just issued.",
     "1. Close the reveal dialog.\n2. Reload list.",
     "—",
     "PIN no longer visible. Only metadata (name, purpose, validity, status) shown."),
    ("Visitor Access", "Revoke", "High",
     "Admin can revoke an active pass",
     "Active visitor pass exists.",
     "1. Click Revoke on a pass row.\n2. Confirm.",
     "—",
     "Pass status changes to revoked. PDK holder disabled / removed from group. PIN no longer opens gate."),
    ("Visitor Access", "Expiration cron", "High",
     "Cron expires passes past validUntil",
     "A visitor pass with validUntil 2 minutes in the past exists.",
     "1. Wait for the next cron run (every ~2 minutes) OR manually trigger the cron endpoint.",
     "—",
     "Pass moves to expired status. expiredAt timestamp set. PDK holder removed."),
    ("Visitor Access", "Filter", "Low",
     "Admin can filter list by status",
     "Mix of active, expired, revoked passes.",
     "1. On /admin/visitor-access change filter to 'Expired'.",
     "—",
     "Only expired passes shown."),
    ("Visitor Access", "Audit", "Medium",
     "Pass entries record createdBy / revokedBy",
     "Issued and revoked passes by known admins.",
     "1. Inspect a pass detail.",
     "—",
     "createdBy = the admin who issued. revokedBy + revokedAt set for revoked passes."),
]

# ===== TEXT-TO-OPEN (Twilio SMS) =====
TC += [
    ("Text-to-Open SMS", "Whitelist authorised", "Critical",
     "Authorised phone unlocks gate via SMS",
     "Twilio webhook URL configured on PDK gate settings. Tester's mobile number added to whitelist.",
     "1. From whitelisted phone, text ANY message (e.g., 'open') to the facility's Twilio number.",
     "Any short text",
     "Reply received: 'Gate opening.' Within seconds entry devices momentarily unlock (verify physically or in PDK panel). AccessLog entry created with source = 'app' or 'sms' and the phone + device IDs."),
    ("Text-to-Open SMS", "Whitelist rejection", "Critical",
     "Non-whitelisted phone is silently rejected",
     "Tester's mobile not on whitelist.",
     "1. From non-whitelisted phone, text the facility number.",
     "Any text",
     "Reply is a generic non-leaking message (or no acknowledgement). NO gate unlock occurs. Console / server logs show rejection. No AccessLog row created."),
    ("Text-to-Open SMS", "Signature verification", "High",
     "Webhook rejects requests with invalid Twilio signature",
     "Tester has access to a webhook tester tool (e.g., curl).",
     "1. POST to /api/webhooks/twilio/sms with bogus X-Twilio-Signature.",
     "Any payload",
     "Response 401/403. No unlock. No DB write."),
    ("Text-to-Open SMS", "Multi-device fan-out", "High",
     "All configured entry devices unlock momentarily",
     "Multiple entry devices configured.",
     "1. Authorised SMS as above.",
     "—",
     "Each entry device receives the momentary unlock command (verify in PDK panel or AccessLog entries: one row per device unlocked)."),
    ("Text-to-Open SMS", "Whitelist edit", "Medium",
     "Admin can add/remove numbers in the whitelist",
     "On /admin/settings/gate.",
     "1. Edit whitelist textarea, add a new number, save.\n2. Reload page.",
     "+15551234567",
     "Saved value is normalised (digits + leading +). Reload shows the updated list."),
]

# ===== PDK GATE INTEGRATION =====
TC += [
    ("PDK Gate", "Kill switch", "Critical",
     "All PDK sync paths short-circuit when kill switch is off",
     "PDK_SYNC_ENABLED unset.",
     "1. Issue / revoke a tenant change.\n2. Inspect logs.",
     "—",
     "No outbound HTTP to PDK. Local DB write succeeds. No errors raised."),
    ("PDK Gate", "Holder sync", "Critical",
     "Tenant changes sync to PDK when enabled",
     "PDK_SYNC_ENABLED=true with valid credentials.",
     "1. Regenerate a tenant gate code.",
     "—",
     "PDK holder PIN updated within seconds. Tenant.pdkSyncedAt updated."),
    ("PDK Gate", "Group rules sync", "High",
     "Editing access hours updates PDK group rules",
     "—",
     "1. Settings → Facility → change access hours.\n2. Save.",
     "Hours 06:00–22:00",
     "PDK group entry rule schedule updated to new hours. Exit rule remains 24/7."),
    ("PDK Gate", "Reconcile job", "Medium",
     "Reconcile cron repairs drift",
     "Manually disable a tenant in PDK panel.",
     "1. Trigger /api/cron/pdkReconcile.",
     "—",
     "Job re-enables tenant and re-adds them to entry group. Summary shows {patched: 1}."),
    ("PDK Gate", "Webhook — allowed", "Critical",
     "Allowed-entry webhook creates AccessLog row",
     "Webhook URL registered, secret configured.",
     "1. Trigger a successful gate entry (real device or simulator).",
     "—",
     "AccessLog row created with tenant attribution, device, timestamp, result=allowed."),
    ("PDK Gate", "Webhook — denied", "High",
     "Denied entry creates denied AccessLog row",
     "—",
     "1. Use a disabled credential at a gate.",
     "—",
     "AccessLog row created with result=denied."),
    ("PDK Gate", "Webhook — visitor", "High",
     "Visitor PIN entry attributes to visitor pass",
     "Active visitor pass with known PIN.",
     "1. Enter visitor PIN at a reader.",
     "—",
     "AccessLog row created with visitorAccessId set (tenantId null). Visible in gate-activity report under visitor."),
]

# ===== STRIPE PAYMENT EDGE CASES =====
TC += [
    ("Stripe / Payments", "Autopay charge", "Critical",
     "Autopay cron charges tenant on billing day",
     "Tenant with autopayEnabled=true, valid card, lease billingDay = today.",
     "1. Trigger autopay cron OR wait for natural run.",
     "—",
     "Payment processed (status=succeeded). Receipt email sent. Balance reduces. Charge appears in payments report."),
    ("Stripe / Payments", "Autopay failure", "High",
     "Failed autopay sends failure notification",
     "Tenant with card that triggers failure on future charge.",
     "1. Trigger autopay cron when due.",
     "Card 4000 0000 0000 0341",
     "Payment fails. Status=failed. Failed-payment email sent to tenant. Admin sees the failed entry."),
    ("Stripe / Payments", "Refund partial", "Medium",
     "Partial refund updates balance correctly",
     "$50 successful payment exists.",
     "1. Refund $20 of the $50 payment.",
     "—",
     "Stripe refund of $20 issued. Tenant balance reflects the partial refund."),
    ("Stripe / Payments", "Late fee", "High",
     "Late fee is added after configured days past due",
     "Delinquent tenant with unpaid invoice older than lateFeeAfterDays.",
     "1. Run delinquency cron.",
     "—",
     "Late fee charge appears in tenant billing history with type=late_fee. Late notice email sent (per template config)."),
    ("Stripe / Payments", "Reservation fee charged at move-in", "Medium",
     "Configured reservation fee is included at move-in",
     "Unit type has reservation fee > 0.",
     "1. Complete public move-in.",
     "—",
     "Reservation fee line item appears on first invoice (or charged at submission, per setting)."),
    ("Stripe / Payments", "Multi-unit payment", "Medium",
     "Tenant with multiple leases can pay all in one checkout",
     "Tenant has 2 active leases with balances.",
     "1. /portal/payments/new — select pay multi.\n2. Submit.",
     "—",
     "Single Stripe charge processes total. Allocations recorded per lease."),
]

# ===== NOTIFICATIONS =====
TC += [
    ("Notifications", "Welcome email", "High",
     "New tenant receives welcome email with credentials",
     "Just completed move-in.",
     "1. Check the tester's inbox after move-in.",
     "—",
     "Welcome email arrives with portal URL and login credentials."),
    ("Notifications", "Payment receipt", "High",
     "Successful payment triggers receipt email",
     "Tenant just made a payment.",
     "1. Check inbox.",
     "—",
     "Receipt email arrives with amount, date, and payment method last 4."),
    ("Notifications", "Late notice", "High",
     "Late notice sent on first overdue day",
     "Tenant with unpaid invoice past lateFeeAfterDays.",
     "1. Run delinquency cron.",
     "—",
     "Late notice email (and SMS if enabled, and print if enabled) sent. Logged in /admin/communications/templates send history."),
    ("Notifications", "Rate change", "Medium",
     "Tenant receives rate-change notification on batch submit",
     "Admin submits a rate change batch including tenant.",
     "1. Submit batch with email channel.",
     "—",
     "Tenant inbox receives rate change notice with effective date."),
    ("Notifications", "Suppression window", "Low",
     "Notifications are suppressed during configured holiday window",
     "Settings configured to suppress 1 day window covering today.",
     "1. Run delinquency cron during suppression window.",
     "—",
     "No notifications go out for that window. Resumes after window."),
    ("Notifications", "SMS opt-out", "Medium",
     "Tenant opt-out blocks SMS notifications",
     "Tenant with smsOptIn=false.",
     "1. Trigger an SMS-channel notification for the tenant.",
     "—",
     "Email is sent. SMS is NOT sent."),
]

# ===== WAITING LIST =====
TC += [
    ("Waiting List", "Admin view", "Medium",
     "Waiting list entries are visible to admin",
     "Public waitlist submission exists.",
     "1. Navigate to /admin/waiting-list.",
     "—",
     "Entries listed with name, requested size, status, join date."),
    ("Waiting List", "Notify", "Medium",
     "Admin can notify a waitlist member of availability",
     "Waitlist entry exists. Matching unit available.",
     "1. Open entry → Notify.",
     "—",
     "Notification email is sent. Entry status changes to notified."),
    ("Waiting List", "Convert", "Medium",
     "Admin can convert waitlist entry to a tenant",
     "Waitlist entry exists.",
     "1. Open entry → Convert.\n2. Fill remaining fields, assign unit, submit.",
     "—",
     "Tenant + lease created from the waitlist data. Entry status = converted."),
]

# ===== MOVE-OUT (admin & tenant flow) =====
TC += [
    ("Move-Out", "Tenant request", "High",
     "Tenant submits move-out request",
     "Tenant logged in with active lease.",
     "1. /portal/move-out → pick lease, date, submit.",
     "Move-out 7 days out",
     "Request submitted. Admin notified. Status pending_moveout."),
    ("Move-Out", "Final settlement", "Critical",
     "Admin can finalize and compute final balance",
     "Pending move-out exists.",
     "1. /admin/move-out → open tenant → review charges → Finalize.",
     "—",
     "Final balance calculated. Receipt generated and emailed. Unit returns to available. Tenant.gate access revoked."),
    ("Move-Out", "Photo upload", "Low",
     "Tenant can upload photos during move-out",
     "On portal move-out screen.",
     "1. Upload 2 photos.\n2. Continue.",
     "Any small JPGs",
     "Upload succeeds. Files visible on admin move-out page."),
    ("Move-Out", "Refund owed", "Medium",
     "Refund-owed move-out routes admin to refund flow",
     "Move-out resulting in negative balance.",
     "1. Finalize a move-out with credit > charges.",
     "—",
     "Final receipt shows refund owed. Admin can issue refund via standard refund flow."),
]

# ===== RETAIL =====
TC += [
    ("Retail", "Walk-in sale (cash)", "Medium",
     "Admin can ring up a walk-in sale to a non-tenant",
     "At least one product with stock.",
     "1. /admin/retail → New walk-in sale → select product, qty, cash → submit.",
     "1x product",
     "Sale recorded. Inventory decremented. Receipt generated."),
    ("Retail", "Walk-in sale (card)", "Medium",
     "Admin can ring up a walk-in sale to a non-tenant via card",
     "Stripe configured.",
     "1. Walk-in sale → card → enter test card → submit.",
     "Test card 4242 4242 4242 4242",
     "Charge succeeds. Receipt generated. Sale appears in retail report."),
    ("Retail", "Walk-in to tenant", "Medium",
     "Sale to existing tenant adds to their balance",
     "Tenant with active account.",
     "1. Walk-in tenant sale → pick tenant → product → submit.",
     "Tenant: john@example.com, qty 1",
     "Charge appears on tenant balance. Inventory decremented."),
    ("Retail", "Inventory adjust", "Low",
     "Admin can adjust inventory with reason",
     "Existing product.",
     "1. Adjust inventory → reason 'damage' → -1.",
     "—",
     "Stock decremented. History entry recorded with reason and admin."),
]

# ===== ACCESSIBILITY / RESPONSIVE / SECURITY (smoke) =====
TC += [
    ("Cross-cutting", "Mobile portal", "High",
     "Tenant portal is usable on mobile",
     "Mobile device or DevTools mobile emulation.",
     "1. Open /portal on iPhone-size viewport.\n2. Navigate dashboard, billing, payments, gate code.",
     "—",
     "Layout adapts. All key actions reachable without horizontal scrolling. Buttons are tap-friendly."),
    ("Cross-cutting", "Mobile move-in", "High",
     "Move-in funnel completes on mobile",
     "Mobile viewport.",
     "1. Full move-in flow on mobile.",
     "—",
     "All steps complete. Stripe CardElement works on mobile. Signature pad works with touch."),
    ("Cross-cutting", "Browser support", "Medium",
     "Portal and admin work on Safari and Edge",
     "Safari (Mac/iOS) and Edge (Win).",
     "1. Repeat key flows on each browser.",
     "—",
     "No visual or functional regressions vs Chrome."),
    ("Cross-cutting", "404 handling", "Low",
     "Unknown URLs show a friendly 404",
     "—",
     "1. Open a non-existent URL.",
     "—",
     "404 page is shown with branding and link back to home."),
    ("Cross-cutting", "HTTPS only", "Medium",
     "Site is served over HTTPS",
     "—",
     "1. Open production URL.",
     "—",
     "Browser shows secure lock. HTTP requests redirect to HTTPS."),
    ("Cross-cutting", "Auth on admin APIs", "Critical",
     "Admin APIs reject unauthenticated requests",
     "—",
     "1. Hit /api/admin/tenants without session.",
     "—",
     "401/403 response. No data leaked."),
    ("Cross-cutting", "Role check on admin APIs", "Critical",
     "Tenant role cannot call admin APIs",
     "Tenant session token.",
     "1. Call admin API with tenant session.",
     "—",
     "403 response. No data leaked."),
]

# ----- Write test cases -----
row = 2
n_cols = len(tc_headers)

# Helper to compose Test ID like 'AUTH-001'
prefix_map = {
    "Authentication": "AUTH",
    "Public Site": "PUB",
    "Move-In": "MOV",
    "Tenant Portal": "POR",
    "Admin / Tenants": "ATEN",
    "Admin / Units": "AUNI",
    "Admin / Payments": "APAY",
    "Admin / Delinquency": "ADEL",
    "Admin / Reports": "AREP",
    "Admin / Settings": "ASET",
    "Admin / Communications": "ACOM",
    "Admin / Rate Mgmt": "ARTE",
    "Visitor Access": "VIS",
    "Text-to-Open SMS": "SMS",
    "PDK Gate": "PDK",
    "Stripe / Payments": "PAY",
    "Notifications": "NTF",
    "Waiting List": "WAIT",
    "Move-Out": "MOUT",
    "Retail": "RET",
    "Cross-cutting": "XCUT",
}
counters = {}

for case in TC:
    module, sub, prio, title, pre, steps, data, expected = case
    pfx = prefix_map.get(module, "GEN")
    counters[pfx] = counters.get(pfx, 0) + 1
    tid = f"{pfx}-{counters[pfx]:03d}"
    values = [tid, module, sub, prio, title, pre, steps, data, expected,
              "", "Not Run", "", "", "", "", ""]
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.alignment = left_wrap
        cell.font = body_font
        cell.border = border
        if (row % 2 == 1):
            cell.fill = alt_fill
    # row height
    # estimate height from longest field
    lengths = [len(str(v)) for v in [steps, expected, pre]]
    est = max(lengths) // 60
    ws.row_dimensions[row].height = max(48, min(160, 24 + est * 14))
    row += 1

last_row = row - 1

# Data validation: Status, Severity, Priority
dv_status = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked,N/A"', allow_blank=True)
dv_status.error = "Pick one: Not Run, Pass, Fail, Blocked, N/A"
dv_status.errorTitle = "Invalid status"
ws.add_data_validation(dv_status)
dv_status.add(f"K2:K{last_row}")

dv_sev = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
ws.add_data_validation(dv_sev)
dv_sev.add(f"L2:L{last_row}")

dv_prio = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
ws.add_data_validation(dv_prio)
dv_prio.add(f"D2:D{last_row}")

# Column widths
set_col_widths(ws, [
    10,   # Test ID
    20,   # Module
    18,   # Sub
    10,   # Priority
    40,   # Title
    35,   # Preconditions
    50,   # Steps
    25,   # Data
    50,   # Expected
    35,   # Actual
    12,   # Status
    12,   # Severity
    14,   # Tester
    12,   # Date
    25,   # Comments
    12,   # Defect ID
])

# Add auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{last_row}"

# Conditional formatting on Status column via cell fills (light, by formula)
from openpyxl.formatting.rule import CellIsRule
green_fill = PatternFill("solid", fgColor=GREEN)
red_fill = PatternFill("solid", fgColor=RED)
yellow_fill = PatternFill("solid", fgColor=YELLOW)
grey_fill = PatternFill("solid", fgColor=GREY)
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="equal", formula=['"Pass"'], fill=green_fill))
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="equal", formula=['"Fail"'], fill=red_fill))
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="equal", formula=['"Blocked"'], fill=yellow_fill))
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="equal", formula=['"N/A"'], fill=grey_fill))

TOTAL_TESTS = len(TC)

# ============================================================
# 5. DEFECT LOG
# ============================================================
ws = wb.create_sheet("Defect Log")
ws.freeze_panes = "A2"
defect_headers = [
    "Defect ID", "Linked Test ID", "Module", "Title", "Severity",
    "Status", "Steps to Reproduce", "Expected", "Actual",
    "Environment / Browser", "Reported By", "Reported Date",
    "Assigned To", "Fix Notes", "Verified By", "Verified Date",
]
header_row(ws, 1, defect_headers, height=36)

# Pre-create 30 empty defect rows with formatting + dropdowns
for r in range(2, 32):
    # Defect ID placeholder
    ws.cell(row=r, column=1, value=f"DEF-{r-1:03d}").font = body_bold
    for c in range(1, len(defect_headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.alignment = left_wrap
        cell.font = body_font
        if r % 2 == 1:
            cell.fill = alt_fill
    ws.row_dimensions[r].height = 40

dv_sev_d = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
ws.add_data_validation(dv_sev_d)
dv_sev_d.add("E2:E31")

dv_status_d = DataValidation(type="list", formula1='"Open,In Progress,Fixed,Verified,Closed,Re-Opened,Wont Fix,Accepted Workaround"', allow_blank=True)
ws.add_data_validation(dv_status_d)
dv_status_d.add("F2:F31")

# Conditional formatting on Defect Status
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Open"'], fill=red_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Fixed"'], fill=yellow_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Verified"'], fill=green_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Closed"'], fill=green_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Re-Opened"'], fill=red_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Wont Fix"'], fill=grey_fill))
ws.conditional_formatting.add("F2:F31",
    CellIsRule(operator="equal", formula=['"Accepted Workaround"'], fill=grey_fill))

set_col_widths(ws, [11, 12, 18, 35, 11, 14, 45, 35, 35, 22, 16, 14, 16, 35, 16, 14])
ws.auto_filter.ref = f"A1:{get_column_letter(len(defect_headers))}31"

# ============================================================
# 6. SUMMARY
# ============================================================
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False

ws.merge_cells("B2:F2")
c = ws["B2"]
c.value = "Test Execution Summary"
c.font = title_font
c.fill = navy_fill
c.alignment = center
ws.row_dimensions[2].height = 28

# Counts by status
ws.cell(row=4, column=2, value="Status").font = body_bold
ws.cell(row=4, column=3, value="Count").font = body_bold
ws.cell(row=4, column=4, value="% of Total").font = body_bold
for col in range(2, 5):
    ws.cell(row=4, column=col).fill = section_fill
    ws.cell(row=4, column=col).border = border

# Total tests is the count of test rows on Test Cases tab (rows 2..1+TOTAL_TESTS)
total_range = f"'Test Cases'!K2:K{1 + TOTAL_TESTS}"

statuses = [
    ("Not Run", "Not Run"),
    ("Pass", "Pass"),
    ("Fail", "Fail"),
    ("Blocked", "Blocked"),
    ("N/A", "N/A"),
]
for i, (label, val) in enumerate(statuses):
    r = 5 + i
    ws.cell(row=r, column=2, value=label).font = body_bold
    ws.cell(row=r, column=3, value=f'=COUNTIF({total_range},"{val}")')
    ws.cell(row=r, column=4, value=f'=IFERROR(C{r}/{TOTAL_TESTS},0)').number_format = "0.0%"
    for col in range(2, 5):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).alignment = left_wrap

# Total row
tr = 5 + len(statuses)
ws.cell(row=tr, column=2, value="TOTAL").font = body_bold
ws.cell(row=tr, column=2).fill = section_fill
ws.cell(row=tr, column=3, value=TOTAL_TESTS).font = body_bold
ws.cell(row=tr, column=3).fill = section_fill
ws.cell(row=tr, column=4, value=1.0).number_format = "0.0%"
ws.cell(row=tr, column=4).fill = section_fill
ws.cell(row=tr, column=4).font = body_bold
for col in range(2, 5):
    ws.cell(row=tr, column=col).border = border

# Counts by priority
pr = tr + 3
ws.cell(row=pr, column=2, value="Priority").font = body_bold
ws.cell(row=pr, column=3, value="Total").font = body_bold
ws.cell(row=pr, column=4, value="Passed").font = body_bold
ws.cell(row=pr, column=5, value="% Pass").font = body_bold
for col in range(2, 6):
    ws.cell(row=pr, column=col).fill = section_fill
    ws.cell(row=pr, column=col).border = border

prio_range = f"'Test Cases'!D2:D{1 + TOTAL_TESTS}"
for i, p in enumerate(["Critical", "High", "Medium", "Low"]):
    r = pr + 1 + i
    ws.cell(row=r, column=2, value=p).font = body_bold
    ws.cell(row=r, column=3, value=f'=COUNTIF({prio_range},"{p}")')
    ws.cell(row=r, column=4, value=f'=COUNTIFS({prio_range},"{p}",{total_range},"Pass")')
    ws.cell(row=r, column=5, value=f'=IFERROR(D{r}/C{r},0)').number_format = "0.0%"
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = border

# Defects by status
dr = pr + 6 + 2
ws.cell(row=dr, column=2, value="Defect Status").font = body_bold
ws.cell(row=dr, column=3, value="Count").font = body_bold
for col in range(2, 4):
    ws.cell(row=dr, column=col).fill = section_fill
    ws.cell(row=dr, column=col).border = border

defect_range = "'Defect Log'!F2:F31"
for i, s in enumerate(["Open", "In Progress", "Fixed", "Verified", "Closed", "Re-Opened", "Wont Fix", "Accepted Workaround"]):
    r = dr + 1 + i
    ws.cell(row=r, column=2, value=s)
    ws.cell(row=r, column=3, value=f'=COUNTIF({defect_range},"{s}")')
    for col in range(2, 4):
        ws.cell(row=r, column=col).border = border

# Sign-off readiness
sor = dr + 11
ws.merge_cells(start_row=sor, start_column=2, end_row=sor, end_column=5)
c = ws.cell(row=sor, column=2, value="Sign-Off Readiness")
c.font = section_font
c.fill = section_fill

crit_total = f'=COUNTIF({prio_range},"Critical")'
crit_pass = f'=COUNTIFS({prio_range},"Critical",{total_range},"Pass")'
high_total = f'=COUNTIF({prio_range},"High")'
high_pass = f'=COUNTIFS({prio_range},"High",{total_range},"Pass")'
open_crit_high = f'=COUNTIFS(\'Defect Log\'!E2:E31,"Critical",\'Defect Log\'!F2:F31,"Open")+COUNTIFS(\'Defect Log\'!E2:E31,"High",\'Defect Log\'!F2:F31,"Open")'

checks = [
    ("Critical tests passing (must be 100%)", f"=IFERROR({crit_pass[1:]}/{crit_total[1:]},0)", "0.0%"),
    ("High tests passing (must be ≥ 95%)", f"=IFERROR({high_pass[1:]}/{high_total[1:]},0)", "0.0%"),
    ("Open Critical/High defects (must be 0)", open_crit_high, "0"),
]
for i, (label, formula, fmt) in enumerate(checks):
    r = sor + 1 + i
    ws.cell(row=r, column=2, value=label).font = body_bold
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2).alignment = left_wrap
    ws.cell(row=r, column=5, value=formula).number_format = fmt
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = border

set_col_widths(ws, [2, 38, 14, 14, 14, 12])

# ============================================================
# Save
# ============================================================
wb.save(OUT)
print(f"WROTE: {OUT}")
print(f"Total test cases: {TOTAL_TESTS}")
